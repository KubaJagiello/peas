import logging
from enum import Enum
from pathlib import Path

from openpyxl import load_workbook

from peas_app.api.dtos.v1.requests.api_requests import ProductRequest
from peas_app.api.mappers.product_mapper import ProductMapper
from peas_app.database.db_config import DatabaseConfig
from peas_app.database.repositories.product_repository import ProductRepository

FILE_NAME = Path("scripts/resources/LivsmedelsDB_202411300928.xlsx")

logging.basicConfig(level=logging.INFO)


class ColumnName(str, Enum):
    PRODUCT_NAME = "Livsmedelsnamn"
    CALORIES = "Energi (kcal)"
    FAT = "Fett, totalt (g)"
    PROTEINS = "Protein (g)"
    CARBOHYDRATES = "Kolhydrater, tillgängliga (g)"
    SALT = "Salt, NaCl (g)"


def get_indexes_for_headers() -> dict[ColumnName, int]:
    workbook = load_workbook(filename=FILE_NAME)
    sheet = workbook.active
    indexes = {}

    column_names = list(sheet.iter_rows(values_only=True))[2]

    for idx, column_name in enumerate(column_names):
        if column_name in ColumnName:
            indexes[ColumnName(column_name)] = idx

    return indexes


def get_product_requests() -> list[ProductRequest]:
    workbook = load_workbook(filename=FILE_NAME)
    sheet = workbook.active

    indexes = get_indexes_for_headers()

    product_requests: list[ProductRequest] = []
    for row in sheet.iter_rows(values_only=True, min_row=4):
        name = row[indexes[ColumnName.PRODUCT_NAME]]
        proteins = row[indexes[ColumnName.PROTEINS]]
        fats = row[indexes[ColumnName.FAT]]
        carbohydrates = row[indexes[ColumnName.CARBOHYDRATES]]
        sodium = row[indexes[ColumnName.SALT]]

        product = ProductRequest(
            name=name,
            proteins=float(proteins),
            fats=float(fats),
            carbohydrates=float(carbohydrates),
            sodium=float(sodium),
        )
        product_requests.append(product)

    return product_requests


def main():
    db_config = DatabaseConfig()
    product_repository = ProductRepository(db_config.get_session())

    product_requests = get_product_requests()
    logging.info(f"Found {len(product_requests)} product_requests")

    for product_request in product_requests:
        logging.info(f"Add {product_requests}")
        product = ProductMapper.to_product(product_request)
        product_repository.add(product)


if __name__ == "__main__":
    main()
